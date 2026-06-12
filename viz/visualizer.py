#!/usr/bin/env python3
"""
Mini-GRP 可视化模块 (Visualizer)

提供专业的金融数据可视化功能，包括:
- Top股票综合得分对比图
- 因子分解堆叠水平条形图
- 行业分布热力图
- 各因子分布直方图
- Excel报告生成 (带条件格式)

配色方案 (低饱和度金融风格):
- 主色: #2C3E50 (深蓝灰)
- 辅色: #34495E (中蓝灰)
- 强调色: #3498DB (蓝色)
- 点缀色: #1ABC9C (青绿)
"""

import os
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# =============================================================================
# Matplotlib 全局配置
# =============================================================================

# 设置中文字体支持
# 优先使用系统可用的CJK字体
cjk_fonts = [
    'Noto Sans CJK SC',
    'WenQuanYi Zen Hei',
    'SimHei',
    'DejaVu Sans',
    'Arial Unicode MS',
]
matplotlib.rcParams['font.sans-serif'] = cjk_fonts
matplotlib.rcParams['axes.unicode_minus'] = False

# 尝试注册字体文件（如果matplotlib无法自动找到）
try:
    import matplotlib.font_manager as fm
    font_paths = [
        '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc',
        '/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc',
    ]
    for fp in font_paths:
        if os.path.exists(fp):
            fm.fontManager.addfont(fp)
except Exception:
    pass

# GRP 配色方案 (低饱和度)
COLOR_PALETTE = {
    'primary': '#2C3E50',       # 深蓝灰 - 主色
    'secondary': '#34495E',     # 中蓝灰 - 辅色
    'accent': '#3498DB',        # 蓝色 - 强调
    'highlight': '#1ABC9C',     # 青绿 - 点缀
    'light': '#ECF0F1',         # 浅灰 - 背景
    'mid': '#95A5A6',           # 中灰 - 网格线
    'dark': '#1A252F',          # 深色 - 标题
    'warm': '#E67E22',          # 橙色 - 辅助
    'purple': '#9B59B6',        # 紫色 - 辅助
}

# 维度颜色映射
DIMENSION_COLORS = {
    'value_score': '#2C3E50',      # 价值 - 深蓝灰
    'quality_score': '#3498DB',     # 质量 - 蓝色
    'growth_score': '#1ABC9C',      # 增长 - 青绿
    'momentum_score': '#E67E22',    # 动量 - 橙色
}

DIMENSION_LABELS = {
    'value_score': '价值',
    'quality_score': '质量',
    'growth_score': '增长',
    'momentum_score': '动量',
}


def _setup_figure(figsize=(12, 8), title=None):
    """
    统一设置图表样式

    Parameters
    ----------
    figsize : tuple
        图表尺寸
    title : str, optional
        图表标题

    Returns
    -------
    fig, ax
        matplotlib 图表对象和坐标轴
    """
    fig, ax = plt.subplots(figsize=figsize)

    # 设置背景色
    fig.patch.set_facecolor('white')
    ax.set_facecolor('#FAFAFA')

    # 设置标题
    if title:
        ax.set_title(title, fontsize=14, fontweight='bold',
                     color=COLOR_PALETTE['dark'], pad=15)

    # 设置网格线
    ax.grid(True, linestyle='--', alpha=0.3, color=COLOR_PALETTE['mid'])
    ax.set_axisbelow(True)

    # 设置边框
    for spine in ax.spines.values():
        spine.set_color(COLOR_PALETTE['mid'])
        spine.set_linewidth(0.5)

    return fig, ax


# =============================================================================
# 图表 1: Top股票综合得分对比图
# =============================================================================

def plot_top_stocks(top_picks: pd.DataFrame, output_path: str):
    """
    绘制Top股票综合得分对比图

    水平条形图，显示综合得分和各维度分解。
    每只股票显示4个维度的堆叠贡献。

    Parameters
    ----------
    top_picks : pd.DataFrame
        Top股票数据，需要包含 value_score, quality_score, growth_score, momentum_score
    output_path : str
        图表保存路径 (.png)
    """
    n = min(len(top_picks), 20)
    df = top_picks.head(n).copy()

    # 反转顺序使得最高分在最上面
    df = df.iloc[::-1]

    fig, ax = _setup_figure(figsize=(12, max(6, n * 0.4)),
                            title=f'Top {n} 股票综合得分对比')

    y_pos = np.arange(len(df))
    bar_height = 0.6

    # 计算每个维度的原始贡献（去标准化后用于显示）
    # 将z-score转换为正数用于堆叠显示
    dim_cols = ['value_score', 'quality_score', 'growth_score', 'momentum_score']

    # 使用标准化后的值进行堆叠显示，需要做min-max归一化到正值
    display_values = {}
    for col in dim_cols:
        values = df[col].values
        # 将z-score映射到 1-10 的范围用于可视化
        vmin, vmax = values.min(), values.max()
        if vmax > vmin:
            display_values[col] = 5 + 5 * (values - vmin) / (vmax - vmin)
        else:
            display_values[col] = np.ones_like(values) * 5

    # 绘制堆叠水平条形图
    left = np.zeros(len(df))
    for col in dim_cols:
        color = DIMENSION_COLORS.get(col, '#2C3E50')
        label = DIMENSION_LABELS.get(col, col)
        values = display_values[col]
        ax.barh(y_pos, values, left=left, height=bar_height,
                color=color, label=label, alpha=0.85, edgecolor='white', linewidth=0.5)
        left += values

    # 设置Y轴标签（股票名称+行业）
    labels = [f"{row['name']} ({row['sw_industry_name']})"
              for _, row in df.iterrows()]
    ax.set_yticks(y_pos)
    ax.set_yticklabels(labels, fontsize=9)

    # 设置X轴
    ax.set_xlabel('维度贡献 (相对值)', fontsize=10, color=COLOR_PALETTE['dark'])

    # 在每个条形右侧显示综合得分
    for i, (_, row) in enumerate(df.iterrows()):
        score = row.get('composite_score', 0)
        ax.text(left[i] + 0.3, i, f'{score:.1f}',
                va='center', fontsize=9, fontweight='bold', color=COLOR_PALETTE['dark'])

    # 图例
    ax.legend(loc='lower right', framealpha=0.9, fontsize=9,
              facecolor='white', edgecolor=COLOR_PALETTE['mid'])

    plt.tight_layout()
    fig.savefig(output_path, dpi=150, bbox_inches='tight',
                facecolor='white', edgecolor='none')
    plt.close(fig)
    print(f"[visualizer] 图表已保存: {output_path}")


# =============================================================================
# 图表 2: 因子分解堆叠水平条形图
# =============================================================================

def plot_factor_breakdown(top_picks: pd.DataFrame, output_path: str):
    """
    绘制因子分解堆叠水平条形图

    显示每只股票四个维度的贡献（使用原始z-score值），
    正贡献向右，负贡献向左，直观展示各维度的优劣。

    Parameters
    ----------
    top_picks : pd.DataFrame
        Top股票数据
    output_path : str
        图表保存路径 (.png)
    """
    n = min(len(top_picks), 20)
    df = top_picks.head(n).copy()
    df = df.iloc[::-1]  # 反转顺序

    fig, ax = _setup_figure(figsize=(12, max(6, n * 0.4)),
                            title=f'Top {n} 股票因子分解')

    y_pos = np.arange(len(df))
    bar_height = 0.6
    dim_cols = ['value_score', 'quality_score', 'growth_score', 'momentum_score']

    # 绘制每个维度的条形（正负分开显示）
    # 正值向右，负值向左
    for idx, col in enumerate(dim_cols):
        color = DIMENSION_COLORS.get(col, '#2C3E50')
        label = DIMENSION_LABELS.get(col, col)
        values = df[col].values

        # 偏移位置以避免重叠
        offset = (idx - 1.5) * (bar_height / 4)

        # 绘制正值（向右）
        positive = np.where(values >= 0, values, 0)
        if np.any(positive > 0):
            ax.barh(y_pos + offset, positive, height=bar_height/4.5,
                    color=color, label=label if idx == 0 else "",
                    alpha=0.85, edgecolor='white', linewidth=0.3)

        # 绘制负值（向左）
        negative = np.where(values < 0, values, 0)
        if np.any(negative < 0):
            ax.barh(y_pos + offset, negative, height=bar_height/4.5,
                    color=color, alpha=0.5, edgecolor='white', linewidth=0.3)

    # 添加零线
    ax.axvline(x=0, color=COLOR_PALETTE['dark'], linewidth=0.8, linestyle='-')

    # 设置Y轴
    labels = [f"{row['name']}" for _, row in df.iterrows()]
    ax.set_yticks(y_pos)
    ax.set_yticklabels(labels, fontsize=9)

    # 设置X轴
    ax.set_xlabel('标准化得分 (z-score)', fontsize=10, color=COLOR_PALETTE['dark'])

    # 添加维度标注
    for idx, col in enumerate(dim_cols):
        offset = (idx - 1.5) * (bar_height / 4)
        color = DIMENSION_COLORS.get(col, '#2C3E50')
        label = DIMENSION_LABELS.get(col, col)
        # 在右侧添加图例文字
        ax.text(ax.get_xlim()[1] * 1.02, len(df) + 0.5 + idx * 0.4,
                f'■ {label}', color=color, fontsize=9, fontweight='bold')

    # 自定义图例
    legend_patches = [mpatches.Patch(color=DIMENSION_COLORS.get(c, '#2C3E50'),
                                      label=DIMENSION_LABELS.get(c, c))
                      for c in dim_cols]
    ax.legend(handles=legend_patches, loc='lower left',
              framealpha=0.9, fontsize=9,
              facecolor='white', edgecolor=COLOR_PALETTE['mid'])

    plt.tight_layout()
    fig.savefig(output_path, dpi=150, bbox_inches='tight',
                facecolor='white', edgecolor='none')
    plt.close(fig)
    print(f"[visualizer] 图表已保存: {output_path}")


# =============================================================================
# 图表 3: 行业分布热力图
# =============================================================================

def plot_industry_distribution(scored_df: pd.DataFrame, output_path: str):
    """
    绘制行业分布热力图

    X轴: 行业
    Y轴: 评分维度 (价值、质量、增长、动量、综合)
    颜色: 平均得分（红色=低分，绿色=高分）

    Parameters
    ----------
    scored_df : pd.DataFrame
        完整评分数据，包含行业分类和各维度得分
    output_path : str
        图表保存路径 (.png)
    """
    if 'sw_industry_name' not in scored_df.columns:
        print("[visualizer] 警告: 缺少 sw_industry_name 列，跳过行业热力图")
        return

    # 准备数据: 各行业各维度的平均得分
    dim_cols = ['value_score', 'quality_score', 'growth_score', 'momentum_score']

    # 计算各行业平均值
    industry_data = scored_df.groupby('sw_industry_name')[dim_cols].mean()

    # 添加综合得分
    if 'composite_score' in scored_df.columns:
        industry_data['composite_score'] = (
            scored_df.groupby('sw_industry_name')['composite_score'].mean()
        )

    # 重命名列
    industry_data = industry_data.rename(columns=DIMENSION_LABELS)
    if 'composite_score' in industry_data.columns:
        industry_data = industry_data.rename(columns={'composite_score': '综合'})

    # 转换为热力图需要的格式 (行为维度，列为行业)
    heatmap_data = industry_data.T

    # 创建图表
    fig, ax = plt.subplots(figsize=(max(10, len(heatmap_data.columns) * 1.2), 6))
    fig.patch.set_facecolor('white')

    # 绘制热力图
    # 使用自定义颜色映射: 红色(低) -> 白色(中) -> 绿色(高)
    cmap = matplotlib.colors.LinearSegmentedColormap.from_list(
        'grp', ['#E74C3C', '#ECF0F1', '#27AE60'], N=100
    )

    im = ax.imshow(heatmap_data.values, cmap=cmap, aspect='auto',
                   vmin=-1, vmax=1)

    # 设置轴标签
    ax.set_xticks(np.arange(len(heatmap_data.columns)))
    ax.set_yticks(np.arange(len(heatmap_data.index)))
    ax.set_xticklabels(heatmap_data.columns, fontsize=10, rotation=45, ha='right')
    ax.set_yticklabels(heatmap_data.index, fontsize=11, fontweight='bold')

    # 在每个单元格中显示数值
    for i in range(len(heatmap_data.index)):
        for j in range(len(heatmap_data.columns)):
            value = heatmap_data.values[i, j]
            text_color = 'white' if abs(value) > 0.5 else COLOR_PALETTE['dark']
            ax.text(j, i, f'{value:.2f}', ha='center', va='center',
                    fontsize=9, color=text_color, fontweight='bold')

    # 添加颜色条
    cbar = plt.colorbar(im, ax=ax, shrink=0.8, pad=0.02)
    cbar.set_label('平均得分', fontsize=10)

    ax.set_title('行业维度得分热力图', fontsize=14, fontweight='bold',
                 color=COLOR_PALETTE['dark'], pad=15)

    plt.tight_layout()
    fig.savefig(output_path, dpi=150, bbox_inches='tight',
                facecolor='white', edgecolor='none')
    plt.close(fig)
    print(f"[visualizer] 图表已保存: {output_path}")


# =============================================================================
# 图表 4: 各因子分布直方图
# =============================================================================

def plot_factor_distribution(scored_df: pd.DataFrame, output_path: str):
    """
    绘制各因子分布直方图

    2x2子图: 价值、质量、增长、动量的分布直方图
    每个子图显示该维度下所有因子的分布

    Parameters
    ----------
    scored_df : pd.DataFrame
        完整评分数据，包含因子列和维度得分列
    output_path : str
        图表保存路径 (.png)
    """
    # 定义每个维度的因子列
    dimension_factors = {
        '价值': ['pe_ttm_factor', 'pb_lf_factor', 'ps_ttm_factor',
                 'ev_ebitda_factor', 'dividend_yield_factor'],
        '质量': ['roe_deducted_factor', 'roa_factor', 'gross_margin_factor',
                 'net_margin_factor', 'debt_to_equity_factor'],
        '增长': ['revenue_yoy_factor', 'profit_yoy_factor'],
        '动量': ['return_1m_factor', 'return_3m_factor', 'return_12m_factor'],
    }

    # 颜色映射
    dim_colors = {
        '价值': '#2C3E50',
        '质量': '#3498DB',
        '增长': '#1ABC9C',
        '动量': '#E67E22',
    }

    fig, axes = plt.subplots(2, 2, figsize=(14, 10))
    fig.patch.set_facecolor('white')
    fig.suptitle('各维度因子分布', fontsize=16, fontweight='bold',
                 color=COLOR_PALETTE['dark'], y=0.98)

    axes = axes.flatten()

    for idx, (dim_name, factors) in enumerate(dimension_factors.items()):
        ax = axes[idx]
        ax.set_facecolor('#FAFAFA')

        color = dim_colors.get(dim_name, '#2C3E50')

        # 过滤掉不存在的列
        available_factors = [f for f in factors if f in scored_df.columns]

        if not available_factors:
            ax.text(0.5, 0.5, f'无{dim_name}因子数据',
                    ha='center', va='center', transform=ax.transAxes,
                    fontsize=12, color=COLOR_PALETTE['mid'])
            ax.set_title(f'{dim_name}因子', fontsize=12, fontweight='bold')
            continue

        # 绘制每个因子的分布
        for f_idx, factor in enumerate(available_factors):
            values = scored_df[factor].dropna()
            if len(values) > 0:
                # 使用半透明叠加直方图
                ax.hist(values, bins=25, alpha=0.4, color=color,
                        edgecolor='white', linewidth=0.3)

        # 如果有维度得分列，也叠加显示
        score_col = f"{dim_name.lower().replace('价值', 'value').replace('质量', 'quality').replace('增长', 'growth').replace('动量', 'momentum')}_score"
        if score_col in scored_df.columns:
            values = scored_df[score_col].dropna()
            if len(values) > 0:
                ax.hist(values, bins=25, alpha=0.7, color=color,
                        edgecolor='white', linewidth=0.5, label=f'{dim_name}综合')

        # 添加均值线
        if available_factors:
            all_values = scored_df[available_factors].values.flatten()
            all_values = all_values[~np.isnan(all_values)]
            if len(all_values) > 0:
                mean_val = np.mean(all_values)
                ax.axvline(mean_val, color=COLOR_PALETTE['warm'],
                           linewidth=1.5, linestyle='--', label=f'均值={mean_val:.2f}')

        # 样式设置
        ax.set_title(f'{dim_name}因子分布', fontsize=12, fontweight='bold',
                     color=COLOR_PALETTE['dark'])
        ax.set_xlabel('标准化得分 (z-score)', fontsize=9)
        ax.set_ylabel('频数', fontsize=9)
        ax.grid(True, linestyle='--', alpha=0.3, color=COLOR_PALETTE['mid'])
        ax.set_axisbelow(True)

        for spine in ax.spines.values():
            spine.set_color(COLOR_PALETTE['mid'])
            spine.set_linewidth(0.5)

        if idx == 0:
            ax.legend(fontsize=8, loc='upper right')

    plt.tight_layout(rect=[0, 0, 1, 0.95])
    fig.savefig(output_path, dpi=150, bbox_inches='tight',
                facecolor='white', edgecolor='none')
    plt.close(fig)
    print(f"[visualizer] 图表已保存: {output_path}")


# =============================================================================
# Excel 报告生成
# =============================================================================

def _create_excel_styles():
    """
    创建Excel样式对象

    Returns
    -------
    dict
        包含各种样式对象的字典
    """
    # 表头样式: 深色背景，白色加粗文字
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 数据样式
    data_font = Font(name='Arial', size=10)
    data_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')

    # 边框
    thin_border = Border(
        left=Side(style='thin', color='BDC3C7'),
        right=Side(style='thin', color='BDC3C7'),
        top=Side(style='thin', color='BDC3C7'),
        bottom=Side(style='thin', color='BDC3C7')
    )

    # 条件格式颜色
    green_fill = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
    light_green_fill = PatternFill(start_color='82E0AA', end_color='82E0AA', fill_type='solid')
    yellow_fill = PatternFill(start_color='F9E79F', end_color='F9E79F', fill_type='solid')
    light_red_fill = PatternFill(start_color='F1948A', end_color='F1948A', fill_type='solid')
    red_fill = PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')
    white_font = Font(color='FFFFFF', bold=True)

    return {
        'header_font': header_font,
        'header_fill': header_fill,
        'header_alignment': header_alignment,
        'data_font': data_font,
        'data_alignment': data_alignment,
        'left_alignment': left_alignment,
        'thin_border': thin_border,
        'green_fill': green_fill,
        'light_green_fill': light_green_fill,
        'yellow_fill': yellow_fill,
        'light_red_fill': light_red_fill,
        'red_fill': red_fill,
        'white_font': white_font,
    }


def _apply_score_conditional_formatting(ws, col_letter, start_row, end_row):
    """
    为得分列应用条件格式（高分绿色，低分红色）

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        工作表对象
    col_letter : str
        列字母
    start_row : int
        起始行号
    end_row : int
        结束行号
    """
    styles = _create_excel_styles()

    for row in range(start_row, end_row + 1):
        cell = ws[f'{col_letter}{row}']
        try:
            value = float(cell.value) if cell.value is not None else 50
        except (ValueError, TypeError):
            value = 50

        # 根据得分设置背景色
        if value >= 80:
            cell.fill = styles['green_fill']
            cell.font = styles['white_font']
        elif value >= 60:
            cell.fill = styles['light_green_fill']
        elif value >= 40:
            cell.fill = styles['yellow_fill']
        elif value >= 20:
            cell.fill = styles['light_red_fill']
        else:
            cell.fill = styles['red_fill']
            cell.font = styles['white_font']


def _write_dataframe_to_sheet(ws, df, start_row=1, start_col=1,
                               apply_score_format=False, score_cols=None):
    """
    将DataFrame写入Excel工作表，并应用格式化

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        工作表对象
    df : pd.DataFrame
        要写入的数据
    start_row : int
        起始行号
    start_col : int
        起始列号
    apply_score_format : bool
        是否对得分列应用条件格式
    score_cols : list
        需要应用条件格式的列名列表
    """
    styles = _create_excel_styles()

    # 写入表头
    for c_idx, col_name in enumerate(df.columns, start=start_col):
        cell = ws.cell(row=start_row, column=c_idx, value=col_name)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['header_alignment']
        cell.border = styles['thin_border']

    # 写入数据
    for r_idx, row in enumerate(df.values, start=start_row + 1):
        for c_idx, value in enumerate(row, start=start_col):
            cell = ws.cell(row=r_idx, column=c_idx)

            # 处理数值格式
            if isinstance(value, (int, float, np.integer, np.floating)):
                if np.isnan(value):
                    cell.value = None
                else:
                    cell.value = round(float(value), 4) if isinstance(value, float) else value
            else:
                cell.value = value

            cell.font = styles['data_font']
            cell.alignment = styles['data_alignment']
            cell.border = styles['thin_border']

    # 应用条件格式
    if apply_score_format and score_cols:
        end_row = start_row + len(df)
        for col_name in score_cols:
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + start_col
                col_letter = get_column_letter(col_idx)
                _apply_score_conditional_formatting(ws, col_letter, start_row + 1, end_row)

    # 设置列宽自适应
    for c_idx, col_name in enumerate(df.columns, start=start_col):
        col_letter = get_column_letter(c_idx)
        max_length = len(str(col_name))
        for value in df[col_name].astype(str):
            max_length = max(max_length, min(len(value), 30))
        adjusted_width = min(max(max_length + 2, 8), 40)
        ws.column_dimensions[col_letter].width = adjusted_width


def _create_top20_sheet(wb, top_picks):
    """
    创建 Top 20 推荐工作表

    Parameters
    ----------
    wb : Workbook
        Excel工作簿
    top_picks : pd.DataFrame
        Top股票数据
    """
    ws = wb.active
    ws.title = 'Top 20 推荐'

    # 准备数据
    display_cols = ['code', 'name', 'sw_industry_name', 'composite_score',
                    'value_score', 'quality_score', 'growth_score', 'momentum_score']

    # 添加industry_rank如果存在
    if 'industry_rank' in top_picks.columns:
        display_cols.append('industry_rank')

    # 确保列都存在
    available_cols = [c for c in display_cols if c in top_picks.columns]
    df = top_picks[available_cols].head(20).copy()

    # 重命名列
    col_name_map = {
        'code': '股票代码',
        'name': '股票名称',
        'sw_industry_name': '行业',
        'composite_score': '综合得分',
        'value_score': '价值得分',
        'quality_score': '质量得分',
        'growth_score': '增长得分',
        'momentum_score': '动量得分',
        'industry_rank': '行业内排名',
    }
    df = df.rename(columns=col_name_map)

    # 写入数据
    score_cols = ['综合得分', '价值得分', '质量得分', '增长得分', '动量得分']
    _write_dataframe_to_sheet(ws, df, start_row=1, start_col=1,
                              apply_score_format=True,
                              score_cols=score_cols)

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 添加标题行
    ws.insert_rows(1)
    title_cell = ws.cell(row=1, column=1, value='Mini-GRP 量化选股系统 - Top 20 推荐')
    title_cell.font = Font(name='Arial', bold=True, size=14, color='2C3E50')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))

    # 调整表头行号（因为插入了标题行）
    for c_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=2, column=c_idx)
        styles = _create_excel_styles()
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['header_alignment']
        cell.border = styles['thin_border']


def _create_full_ranking_sheet(wb, scored_df):
    """
    创建全部排名工作表

    Parameters
    ----------
    wb : Workbook
        Excel工作簿
    scored_df : pd.DataFrame
        完整评分数据
    """
    ws = wb.create_sheet(title='全部排名')

    # 选择关键列
    key_cols = ['code', 'name', 'sw_industry_name', 'composite_score',
                'value_score', 'quality_score', 'growth_score', 'momentum_score']
    if 'industry_rank' in scored_df.columns:
        key_cols.append('industry_rank')

    available_cols = [c for c in key_cols if c in scored_df.columns]

    # 按综合得分排序
    df = scored_df[available_cols].sort_values('composite_score', ascending=False).copy()

    # 重命名列
    col_name_map = {
        'code': '股票代码',
        'name': '股票名称',
        'sw_industry_name': '行业',
        'composite_score': '综合得分',
        'value_score': '价值得分',
        'quality_score': '质量得分',
        'growth_score': '增长得分',
        'momentum_score': '动量得分',
        'industry_rank': '行业内排名',
    }
    df = df.rename(columns=col_name_map)

    # 写入数据
    score_cols = ['综合得分', '价值得分', '质量得分', '增长得分', '动量得分']
    _write_dataframe_to_sheet(ws, df, start_row=1, start_col=1,
                              apply_score_format=True,
                              score_cols=score_cols)

    # 冻结首行
    ws.freeze_panes = 'A2'


def _create_industry_stats_sheet(wb, scored_df):
    """
    创建行业统计工作表

    Parameters
    ----------
    wb : Workbook
        Excel工作簿
    scored_df : pd.DataFrame
        完整评分数据
    """
    ws = wb.create_sheet(title='行业统计')

    if 'sw_industry_name' not in scored_df.columns:
        ws.cell(row=1, column=1, value='无行业数据')
        return

    # 计算行业统计
    dim_cols = ['value_score', 'quality_score', 'growth_score', 'momentum_score']
    available_dim_cols = [c for c in dim_cols if c in scored_df.columns]

    industry_stats = scored_df.groupby('sw_industry_name').agg(
        股票数量=('code', 'count'),
        平均综合得分=('composite_score', 'mean'),
        **{f'平均{DIMENSION_LABELS.get(c, c)}得分': (c, 'mean')
           for c in available_dim_cols}
    ).reset_index()

    industry_stats = industry_stats.rename(columns={'sw_industry_name': '行业'})
    industry_stats = industry_stats.sort_values('平均综合得分', ascending=False)

    # 四舍五入
    numeric_cols = industry_stats.select_dtypes(include=[np.number]).columns
    industry_stats[numeric_cols] = industry_stats[numeric_cols].round(4)

    # 写入数据
    score_cols = [c for c in industry_stats.columns if '得分' in c]
    _write_dataframe_to_sheet(ws, industry_stats, start_row=1, start_col=1,
                              apply_score_format=False)

    # 冻结首行
    ws.freeze_panes = 'A2'


# =============================================================================
# 完整报告生成
# =============================================================================

def generate_report(scored_df: pd.DataFrame, top_picks: pd.DataFrame, output_dir: str):
    """
    生成完整报告

    1. 保存4张图表到 output_dir:
       - top20_scores.png: Top20综合得分对比
       - factor_breakdown.png: 因子分解堆叠图
       - industry_heatmap.png: 行业热力图
       - factor_distribution.png: 因子分布图

    2. 保存Excel报告到 output_dir/report.xlsx:
       - Sheet1: Top 20 推荐 (带格式化的表格)
       - Sheet2: 全部排名 (所有股票完整数据)
       - Sheet3: 行业统计 (各行业平均得分)

    Parameters
    ----------
    scored_df : pd.DataFrame
        完整的评分结果数据
    top_picks : pd.DataFrame
        Top N 推荐股票
    output_dir : str
        输出目录路径
    """
    os.makedirs(output_dir, exist_ok=True)

    print("=" * 60)
    print("[visualizer] 开始生成报告")
    print("=" * 60)

    # --- 生成图表 ---
    print("\n[visualizer] 生成图表...")

    # 图1: Top股票综合得分对比
    plot_top_stocks(top_picks, os.path.join(output_dir, 'top20_scores.png'))

    # 图2: 因子分解堆叠图
    plot_factor_breakdown(top_picks, os.path.join(output_dir, 'factor_breakdown.png'))

    # 图3: 行业热力图
    plot_industry_distribution(scored_df, os.path.join(output_dir, 'industry_heatmap.png'))

    # 图4: 因子分布图
    plot_factor_distribution(scored_df, os.path.join(output_dir, 'factor_distribution.png'))

    # --- 生成Excel报告 ---
    print("\n[visualizer] 生成Excel报告...")

    wb = Workbook()

    # Sheet 1: Top 20 推荐
    print("  - Sheet 1: Top 20 推荐")
    _create_top20_sheet(wb, top_picks)

    # Sheet 2: 全部排名
    print("  - Sheet 2: 全部排名")
    _create_full_ranking_sheet(wb, scored_df)

    # Sheet 3: 行业统计
    print("  - Sheet 3: 行业统计")
    _create_industry_stats_sheet(wb, scored_df)

    # 保存Excel
    excel_path = os.path.join(output_dir, 'report.xlsx')
    wb.save(excel_path)
    print(f"  - Excel已保存: {excel_path}")

    print("\n" + "=" * 60)
    print("[visualizer] 报告生成完成!")
    print(f"  - 图表: {output_dir}/{{top20_scores.png, factor_breakdown.png, industry_heatmap.png, factor_distribution.png}}")
    print(f"  - Excel: {excel_path}")
    print("=" * 60)


if __name__ == '__main__':
    # 测试可视化模块
    import sys
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

    from mock_data import generate_mock_data
    from scoring_engine import run_full_scoring, get_top_picks

    print("=" * 60)
    print("测试可视化模块")
    print("=" * 60)

    # 生成模拟数据
    df = generate_mock_data(n_stocks=100)

    # 运行评分
    scored = run_full_scoring(df)
    top20 = get_top_picks(scored, n=20)

    # 生成报告
    output_dir = '/mnt/agents/output/mini_grp/test_output'
    generate_report(scored, top20, output_dir)
